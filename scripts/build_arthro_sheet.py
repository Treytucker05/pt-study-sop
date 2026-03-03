import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import os

def create_cheat_sheet():
    data = [
        # Upper Extremity - Shoulder
        {"Region": "Upper Extremity", "Joint": "Glenohumeral", "Rule": "Convex on Concave", "Motion": "Flexion", "Glide Direction": "Posterior / Inferior", "Resting/Open Pack": "55° Abduction, 30° Horiz Adduction", "Closed Pack": "Full Abduction & External Rotation", "Capsular Pattern": "ER > ABD > IR"},
        {"Region": "Upper Extremity", "Joint": "Glenohumeral", "Rule": "Convex on Concave", "Motion": "Extension", "Glide Direction": "Anterior / Superior", "Resting/Open Pack": "55° Abduction, 30° Horiz Adduction", "Closed Pack": "Full Abduction & External Rotation", "Capsular Pattern": "ER > ABD > IR"},
        {"Region": "Upper Extremity", "Joint": "Glenohumeral", "Rule": "Convex on Concave", "Motion": "Abduction", "Glide Direction": "Inferior", "Resting/Open Pack": "55° Abduction, 30° Horiz Adduction", "Closed Pack": "Full Abduction & External Rotation", "Capsular Pattern": "ER > ABD > IR"},
        {"Region": "Upper Extremity", "Joint": "Glenohumeral", "Rule": "Convex on Concave", "Motion": "External Rotation", "Glide Direction": "Anterior", "Resting/Open Pack": "55° Abduction, 30° Horiz Adduction", "Closed Pack": "Full Abduction & External Rotation", "Capsular Pattern": "ER > ABD > IR"},
        {"Region": "Upper Extremity", "Joint": "Glenohumeral", "Rule": "Convex on Concave", "Motion": "Internal Rotation", "Glide Direction": "Posterior", "Resting/Open Pack": "55° Abduction, 30° Horiz Adduction", "Closed Pack": "Full Abduction & External Rotation", "Capsular Pattern": "ER > ABD > IR"},
        
        # Upper Extremity - Elbow/Forearm
        {"Region": "Upper Extremity", "Joint": "Humeroulnar", "Rule": "Concave on Convex", "Motion": "Flexion", "Glide Direction": "Anterior", "Resting/Open Pack": "70° Flexion, 10° Supination", "Closed Pack": "Full Extension w/ Supination", "Capsular Pattern": "Flexion > Extension"},
        {"Region": "Upper Extremity", "Joint": "Humeroulnar", "Rule": "Concave on Convex", "Motion": "Extension", "Glide Direction": "Posterior", "Resting/Open Pack": "70° Flexion, 10° Supination", "Closed Pack": "Full Extension w/ Supination", "Capsular Pattern": "Flexion > Extension"},
        {"Region": "Upper Extremity", "Joint": "Humeroradial", "Rule": "Concave on Convex", "Motion": "Flexion", "Glide Direction": "Anterior", "Resting/Open Pack": "Full Extension, Full Supination", "Closed Pack": "90° Flexion, 5° Supination", "Capsular Pattern": "Flexion > Extension"},
        {"Region": "Upper Extremity", "Joint": "Humeroradial", "Rule": "Concave on Convex", "Motion": "Extension", "Glide Direction": "Posterior", "Resting/Open Pack": "Full Extension, Full Supination", "Closed Pack": "90° Flexion, 5° Supination", "Capsular Pattern": "Flexion > Extension"},
        {"Region": "Upper Extremity", "Joint": "Proximal Radioulnar", "Rule": "Convex on Concave", "Motion": "Supination", "Glide Direction": "Anterior", "Resting/Open Pack": "70° Flexion, 35° Supination", "Closed Pack": "5° Supination", "Capsular Pattern": "Supination = Pronation"},
        {"Region": "Upper Extremity", "Joint": "Proximal Radioulnar", "Rule": "Convex on Concave", "Motion": "Pronation", "Glide Direction": "Posterior", "Resting/Open Pack": "70° Flexion, 35° Supination", "Closed Pack": "5° Supination", "Capsular Pattern": "Supination = Pronation"},
        {"Region": "Upper Extremity", "Joint": "Distal Radioulnar", "Rule": "Concave on Convex", "Motion": "Supination", "Glide Direction": "Posterior (Dorsal)", "Resting/Open Pack": "10° Supination", "Closed Pack": "5° Supination", "Capsular Pattern": "Supination = Pronation"},
        {"Region": "Upper Extremity", "Joint": "Distal Radioulnar", "Rule": "Concave on Convex", "Motion": "Pronation", "Glide Direction": "Anterior (Volar)", "Resting/Open Pack": "10° Supination", "Closed Pack": "5° Supination", "Capsular Pattern": "Supination = Pronation"},

        # Upper Extremity - Wrist/Hand
        {"Region": "Upper Extremity", "Joint": "Radiocarpal", "Rule": "Convex on Concave", "Motion": "Flexion", "Glide Direction": "Posterior (Dorsal)", "Resting/Open Pack": "Neutral w/ slight Ulnar Dev", "Closed Pack": "Extension w/ Radial Dev", "Capsular Pattern": "Flexion = Extension"},
        {"Region": "Upper Extremity", "Joint": "Radiocarpal", "Rule": "Convex on Concave", "Motion": "Extension", "Glide Direction": "Anterior (Volar)", "Resting/Open Pack": "Neutral w/ slight Ulnar Dev", "Closed Pack": "Extension w/ Radial Dev", "Capsular Pattern": "Flexion = Extension"},
        {"Region": "Upper Extremity", "Joint": "Radiocarpal", "Rule": "Convex on Concave", "Motion": "Radial Deviation", "Glide Direction": "Ulnar", "Resting/Open Pack": "Neutral w/ slight Ulnar Dev", "Closed Pack": "Extension w/ Radial Dev", "Capsular Pattern": "Flexion = Extension"},
        {"Region": "Upper Extremity", "Joint": "Radiocarpal", "Rule": "Convex on Concave", "Motion": "Ulnar Deviation", "Glide Direction": "Radial", "Resting/Open Pack": "Neutral w/ slight Ulnar Dev", "Closed Pack": "Extension w/ Radial Dev", "Capsular Pattern": "Flexion = Extension"},
        {"Region": "Upper Extremity", "Joint": "MCP / IP Joints (Fingers)", "Rule": "Concave on Convex", "Motion": "Flexion", "Glide Direction": "Anterior (Volar)", "Resting/Open Pack": "Slight Flexion", "Closed Pack": "MCP: Full Flexion / IP: Full Ext", "Capsular Pattern": "Flexion > Extension"},
        {"Region": "Upper Extremity", "Joint": "MCP / IP Joints (Fingers)", "Rule": "Concave on Convex", "Motion": "Extension", "Glide Direction": "Posterior (Dorsal)", "Resting/Open Pack": "Slight Flexion", "Closed Pack": "MCP: Full Flexion / IP: Full Ext", "Capsular Pattern": "Flexion > Extension"},

        # Lower Extremity - Hip
        {"Region": "Lower Extremity", "Joint": "Hip / Coxafemoral", "Rule": "Convex on Concave", "Motion": "Flexion", "Glide Direction": "Posterior / Inferior", "Resting/Open Pack": "30° Flexion, 30° Abduction, Slight ER", "Closed Pack": "Full Extension, IR, Abduction", "Capsular Pattern": "Flexion > Abduction > IR"},
        {"Region": "Lower Extremity", "Joint": "Hip / Coxafemoral", "Rule": "Convex on Concave", "Motion": "Extension", "Glide Direction": "Anterior", "Resting/Open Pack": "30° Flexion, 30° Abduction, Slight ER", "Closed Pack": "Full Extension, IR, Abduction", "Capsular Pattern": "Flexion > Abduction > IR"},
        {"Region": "Lower Extremity", "Joint": "Hip / Coxafemoral", "Rule": "Convex on Concave", "Motion": "Abduction", "Glide Direction": "Inferior", "Resting/Open Pack": "30° Flexion, 30° Abduction, Slight ER", "Closed Pack": "Full Extension, IR, Abduction", "Capsular Pattern": "Flexion > Abduction > IR"},
        {"Region": "Lower Extremity", "Joint": "Hip / Coxafemoral", "Rule": "Convex on Concave", "Motion": "Adduction", "Glide Direction": "Superior", "Resting/Open Pack": "30° Flexion, 30° Abduction, Slight ER", "Closed Pack": "Full Extension, IR, Abduction", "Capsular Pattern": "Flexion > Abduction > IR"},
        {"Region": "Lower Extremity", "Joint": "Hip / Coxafemoral", "Rule": "Convex on Concave", "Motion": "Internal Rotation", "Glide Direction": "Posterior", "Resting/Open Pack": "30° Flexion, 30° Abduction, Slight ER", "Closed Pack": "Full Extension, IR, Abduction", "Capsular Pattern": "Flexion > Abduction > IR"},
        {"Region": "Lower Extremity", "Joint": "Hip / Coxafemoral", "Rule": "Convex on Concave", "Motion": "External Rotation", "Glide Direction": "Anterior", "Resting/Open Pack": "30° Flexion, 30° Abduction, Slight ER", "Closed Pack": "Full Extension, IR, Abduction", "Capsular Pattern": "Flexion > Abduction > IR"},

        # Lower Extremity - Knee
        {"Region": "Lower Extremity", "Joint": "Tibiofemoral", "Rule": "Concave on Convex", "Motion": "Flexion", "Glide Direction": "Posterior", "Resting/Open Pack": "25° Flexion", "Closed Pack": "Full Extension, ER of Tibia", "Capsular Pattern": "Flexion > Extension"},
        {"Region": "Lower Extremity", "Joint": "Tibiofemoral", "Rule": "Concave on Convex", "Motion": "Extension", "Glide Direction": "Anterior", "Resting/Open Pack": "25° Flexion", "Closed Pack": "Full Extension, ER of Tibia", "Capsular Pattern": "Flexion > Extension"},
        {"Region": "Lower Extremity", "Joint": "Patellofemoral", "Rule": "N/A", "Motion": "Flexion (Knee)", "Glide Direction": "Inferior", "Resting/Open Pack": "Full Extension", "Closed Pack": "Full Flexion", "Capsular Pattern": "N/A"},
        {"Region": "Lower Extremity", "Joint": "Patellofemoral", "Rule": "N/A", "Motion": "Extension (Knee)", "Glide Direction": "Superior", "Resting/Open Pack": "Full Extension", "Closed Pack": "Full Flexion", "Capsular Pattern": "N/A"},
        
        # Lower Extremity - Ankle/Foot
        {"Region": "Lower Extremity", "Joint": "Talocrural", "Rule": "Convex on Concave", "Motion": "Dorsiflexion", "Glide Direction": "Posterior", "Resting/Open Pack": "10° Plantarflexion", "Closed Pack": "Maximum Dorsiflexion", "Capsular Pattern": "Plantarflexion > Dorsiflexion"},
        {"Region": "Lower Extremity", "Joint": "Talocrural", "Rule": "Convex on Concave", "Motion": "Plantarflexion", "Glide Direction": "Anterior", "Resting/Open Pack": "10° Plantarflexion", "Closed Pack": "Maximum Dorsiflexion", "Capsular Pattern": "Plantarflexion > Dorsiflexion"},
        {"Region": "Lower Extremity", "Joint": "Subtalar", "Rule": "Convex on Concave (Post. Facet)", "Motion": "Inversion", "Glide Direction": "Lateral", "Resting/Open Pack": "Midway between extremes", "Closed Pack": "Supination", "Capsular Pattern": "Varus > Valgus limitation"},
        {"Region": "Lower Extremity", "Joint": "Subtalar", "Rule": "Convex on Concave (Post. Facet)", "Motion": "Eversion", "Glide Direction": "Medial", "Resting/Open Pack": "Midway between extremes", "Closed Pack": "Supination", "Capsular Pattern": "Varus > Valgus limitation"},
        {"Region": "Lower Extremity", "Joint": "MTP / IP Joints (Toes)", "Rule": "Concave on Convex", "Motion": "Flexion", "Glide Direction": "Plantar", "Resting/Open Pack": "Neutral", "Closed Pack": "Full Extension", "Capsular Pattern": "Extension > Flexion"},
        {"Region": "Lower Extremity", "Joint": "MTP / IP Joints (Toes)", "Rule": "Concave on Convex", "Motion": "Extension", "Glide Direction": "Dorsal", "Resting/Open Pack": "Neutral", "Closed Pack": "Full Extension", "Capsular Pattern": "Extension > Flexion"},

        # Spine
        {"Region": "Spine", "Joint": "Cervical Facets (C2-C7)", "Rule": "Plane Joints", "Motion": "Flexion", "Glide Direction": "Superior & Anterior (Bilateral)", "Resting/Open Pack": "Midway between Flex & Ext", "Closed Pack": "Full Extension", "Capsular Pattern": "Lateral Flexion = Rotation, Extension"},
        {"Region": "Spine", "Joint": "Cervical Facets (C2-C7)", "Rule": "Plane Joints", "Motion": "Extension", "Glide Direction": "Inferior & Posterior (Bilateral)", "Resting/Open Pack": "Midway between Flex & Ext", "Closed Pack": "Full Extension", "Capsular Pattern": "Lateral Flexion = Rotation, Extension"},
        {"Region": "Spine", "Joint": "Cervical Facets (C2-C7)", "Rule": "Plane Joints", "Motion": "Right Sidebending", "Glide Direction": "Right: Inf/Post, Left: Sup/Ant", "Resting/Open Pack": "Midway between Flex & Ext", "Closed Pack": "Full Extension", "Capsular Pattern": "Lateral Flexion = Rotation, Extension"},
        {"Region": "Spine", "Joint": "Cervical Facets (C2-C7)", "Rule": "Plane Joints", "Motion": "Right Rotation", "Glide Direction": "Right: Inf/Post, Left: Sup/Ant", "Resting/Open Pack": "Midway between Flex & Ext", "Closed Pack": "Full Extension", "Capsular Pattern": "Lateral Flexion = Rotation, Extension"},
        {"Region": "Spine", "Joint": "Thoracic & Lumbar Facets", "Rule": "Plane Joints", "Motion": "Flexion", "Glide Direction": "Superior & Anterior (Bilateral)", "Resting/Open Pack": "Midway between Flex & Ext", "Closed Pack": "Full Extension", "Capsular Pattern": "Lateral Flexion = Rotation, Extension"},
        {"Region": "Spine", "Joint": "Thoracic & Lumbar Facets", "Rule": "Plane Joints", "Motion": "Extension", "Glide Direction": "Inferior & Posterior (Bilateral)", "Resting/Open Pack": "Midway between Flex & Ext", "Closed Pack": "Full Extension", "Capsular Pattern": "Lateral Flexion = Rotation, Extension"},
    ]

    df = pd.DataFrame(data)
    
    output_file = "arthrokinematics_cheat_sheet.xlsx"
    
    # Save using pandas first to create the base file
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="Arthrokinematics")
    
    # Reload with openpyxl to apply rich formatting
    wb = load_workbook(output_file)
    ws = wb.active
    
    # Formatting Header
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = min(adjusted_width, 40) # Cap width at 40
        
        # Center align data for specific columns
        if column in ['C', 'D', 'E', 'F', 'G', 'H']:
            for cell in col[1:]:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Freeze top row
    ws.freeze_panes = "A2"
    
    wb.save(output_file)
    print(f"Successfully generated {output_file}")

if __name__ == "__main__":
    create_cheat_sheet()